"""
https://openpyxl.readthedocs.io/en/stable

pip install openpyxl
pipenv install openpyxl
"""
import random

import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
# Vendo quantas abas tenho dentro da planilha
nomePlanilhas = pedidos.sheetnames
# print(nomePlanilhas)

planilha1 = pedidos['Página1']
# print(planilha1['b4'].value)  # Vendo o valor que está na coluna B, Linha 4

# Vendo todos dados que estão em uma coluna específica
# for campo in planilha1['b']:
#     if campo.value is None:
#         break
#     print(campo.value)
#
# Acessando ranges na planilha
# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# listando a planilha completa
# contador = 0
# for linha in planilha1:
#     if linha[contador].value is not None:
#         print(linha[contador].value, end=" ")
#
#     contador += 1
#     if linha[contador].value is not None:
#         print(linha[contador].value, end=" ")
#
#     contador += 1
#     if linha[contador].value is not None:
#         print(linha[contador].value)
#
#     contador = 0

# Editando o valor de uma coluna
# Não edita o arquivo original e sim o objeto planilha1 que criamos
# planilha1['b3'] = 2200
# Para salvar o objeto com os valores alterados
# pedidos.save('novaPlanilha.xlsx')

# Adicionando dados a planilha para várias células
# for linha in range(5, 16):
#     numeroPedido = linha - 1
#     planilha1.cell(linha, 1).value = numeroPedido
#     planilha1.cell(linha, 2).value = 1200 + linha
#     preco = round(random.uniform(10, 100))
#     planilha1.cell(linha, 3).value = preco
#
#
# pedidos.save('novaPlanilha.xlsx')

# Criando uma nova aba dentro da planilha
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha2', 0)  # Aba 1
planilha.create_sheet('Planilha3', 1)  # Aba 2

planilha2 = planilha['Planilha2']  # Objeto da planilha 2
planilha3 = planilha['Planilha3']  # Objeto da planilha 3

# Incluindo valores na planilha2
for linha in range(5, 16):
    numeroPedido = linha - 1
    planilha2.cell(linha, 1).value = numeroPedido
    planilha2.cell(linha, 2).value = 1200 + linha
    preco = round(random.uniform(10, 100), 2)
    planilha2.cell(linha, 3).value = preco

# Incluindo valores na planilha3
for linha in range(5, 16):
    planilha3.cell(linha, 1).value = f'Luiz {linha} {round(random.uniform(10, 100), 2)}'
    planilha3.cell(linha, 1).value = f'Otávio {linha} {round(random.uniform(10, 100), 2)}'
    planilha3.cell(linha, 1).value = f'Daniel {linha} {round(random.uniform(10, 100), 2)}'

# Salvando em arquivo
planilha.save('planilhaComAbas.xlsx')
